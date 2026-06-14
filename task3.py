#!/usr/bin/env python3
"""
Security Assessment Toolkit
Dark-themed GUI with port scanning, DNS lookup, HTTP header analysis,
WHOIS lookup, PDF report generation, and Excel export.
"""

import tkinter as tk
from tkinter import ttk, messagebox, scrolledtext
import threading
import socket
import datetime
import os
import sys
import csv
import json

# ─── Color Palette ───────────────────────────────────────────────────────────
BG        = "#0d1117"
BG2       = "#161b22"
BG3       = "#21262d"
ACCENT    = "#58a6ff"
ACCENT2   = "#388bfd"
GREEN     = "#3fb950"
RED       = "#f85149"
YELLOW    = "#d29922"
TEXT      = "#e6edf3"
TEXT_DIM  = "#8b949e"
BORDER    = "#30363d"

COMMON_PORTS = {
    21: "FTP", 22: "SSH", 23: "Telnet", 25: "SMTP",
    53: "DNS", 80: "HTTP", 110: "POP3", 143: "IMAP",
    443: "HTTPS", 445: "SMB", 3306: "MySQL",
    3389: "RDP", 5432: "PostgreSQL", 6379: "Redis",
    8080: "HTTP-Alt", 8443: "HTTPS-Alt", 27017: "MongoDB"
}

# ─── Helper Functions ─────────────────────────────────────────────────────────

def scan_port(host, port, timeout=0.5):
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        s.settimeout(timeout)
        result = s.connect_ex((host, port))
        s.close()
        return result == 0
    except Exception:
        return False

def dns_lookup(host):
    results = {}
    try:
        results["IPv4"] = socket.gethostbyname(host)
    except Exception as e:
        results["IPv4"] = f"Error: {e}"
    try:
        info = socket.getaddrinfo(host, None, socket.AF_INET6)
        results["IPv6"] = info[0][4][0] if info else "Not found"
    except Exception:
        results["IPv6"] = "Not available"
    try:
        hostname, aliases, _ = socket.gethostbyaddr(host)
        results["Hostname"] = hostname
        results["Aliases"] = ", ".join(aliases) if aliases else "None"
    except Exception:
        results["Hostname"] = "Reverse lookup failed"
    return results

def http_headers(host):
    import urllib.request, urllib.error
    results = {}
    for scheme in ["https", "http"]:
        url = f"{scheme}://{host}"
        try:
            req = urllib.request.Request(url, headers={"User-Agent": "SecurityToolkit/1.0"})
            resp = urllib.request.urlopen(req, timeout=5)
            hdrs = dict(resp.headers)
            results[f"{scheme.upper()} Status"] = str(resp.status)
            for k, v in hdrs.items():
                results[f"{scheme.upper()} {k}"] = v
            break
        except urllib.error.HTTPError as e:
            results[f"{scheme.upper()} Status"] = str(e.code)
        except Exception as e:
            results[f"{scheme.upper()} Error"] = str(e)
    return results

def whois_lookup(host):
    """Simple WHOIS via socket (port 43)."""
    results = {}
    try:
        ip = socket.gethostbyname(host)
        results["Resolved IP"] = ip
    except Exception as e:
        results["Error"] = str(e)
        return results
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        s.settimeout(5)
        s.connect(("whois.iana.org", 43))
        s.send((ip + "\r\n").encode())
        raw = b""
        while True:
            chunk = s.recv(4096)
            if not chunk:
                break
            raw += chunk
        s.close()
        lines = raw.decode(errors="replace").splitlines()
        for line in lines[:30]:
            line = line.strip()
            if ":" in line and not line.startswith("%"):
                k, _, v = line.partition(":")
                results[k.strip()] = v.strip()
    except Exception as e:
        results["WHOIS Error"] = str(e)
    return results

# ─── Ping / ICMP Echo ────────────────────────────────────────────────────────

def ping_host(host, count=5):
    """
    Cross-platform ping using raw ICMP (requires root/admin) with fallback
    to TCP-connect latency on port 80/443 so it always works unprivileged.
    Returns list of dicts: {seq, latency_ms, status}
    """
    import time, struct, select

    results = []

    def _tcp_ping(h, p=80, timeout=2.0):
        t0 = time.perf_counter()
        try:
            s = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
            s.settimeout(timeout)
            err = s.connect_ex((h, p))
            rtt = (time.perf_counter() - t0) * 1000
            s.close()
            return rtt if err == 0 else None
        except Exception:
            return None

    def _icmp_ping(ip, seq, timeout=2.0):
        """Send one ICMP echo request; returns RTT ms or None."""
        ICMP_ECHO = 8
        try:
            sock = socket.socket(socket.AF_INET, socket.SOCK_RAW, socket.IPPROTO_ICMP)
            sock.settimeout(timeout)
        except PermissionError:
            return None, "no_raw"

        # Build packet
        pid     = os.getpid() & 0xFFFF
        header  = struct.pack("bbHHh", ICMP_ECHO, 0, 0, pid, seq)
        payload = b"SecurityToolkit " * 2
        chk     = _icmp_checksum(header + payload)
        header  = struct.pack("bbHHh", ICMP_ECHO, 0, chk, pid, seq)
        packet  = header + payload

        t0 = time.perf_counter()
        try:
            sock.sendto(packet, (ip, 1))
            ready = select.select([sock], [], [], timeout)
            if ready[0]:
                _ = sock.recv(1024)
                rtt = (time.perf_counter() - t0) * 1000
                sock.close()
                return rtt, "ok"
            sock.close()
            return None, "timeout"
        except Exception:
            try:
                sock.close()
            except Exception:
                pass
            return None, "error"

    def _icmp_checksum(data):
        s = 0
        for i in range(0, len(data) - 1, 2):
            w = (data[i] << 8) + data[i + 1]
            s += w
        if len(data) % 2:
            s += data[-1] << 8
        s  = (s >> 16) + (s & 0xFFFF)
        s += (s >> 16)
        return ~s & 0xFFFF

    # Resolve host
    try:
        ip = socket.gethostbyname(host)
    except Exception as e:
        return [{"seq": 0, "latency_ms": None, "status": f"DNS error: {e}"}]

    method = "ICMP"
    for seq in range(1, count + 1):
        rtt, flag = _icmp_ping(ip, seq)
        if flag == "no_raw":
            # Fallback: TCP-connect latency
            method = "TCP"
            for port in [80, 443, 22]:
                rtt = _tcp_ping(ip, port)
                if rtt is not None:
                    break
            flag = "ok" if rtt is not None else "timeout"

        results.append({
            "seq":        seq,
            "latency_ms": round(rtt, 2) if rtt is not None else None,
            "status":     "OK" if rtt is not None else "Timeout",
        })

    successful = [r["latency_ms"] for r in results if r["latency_ms"] is not None]
    summary = {
        "host":     host,
        "ip":       ip,
        "method":   method,
        "sent":     count,
        "received": len(successful),
        "lost":     count - len(successful),
        "loss_pct": round((count - len(successful)) / count * 100, 1),
        "min_ms":   round(min(successful), 2) if successful else None,
        "max_ms":   round(max(successful), 2) if successful else None,
        "avg_ms":   round(sum(successful) / len(successful), 2) if successful else None,
        "probes":   results,
    }
    return summary


# ─── Vulnerability Scanner ────────────────────────────────────────────────────

# Known vulnerable server banners / version substrings (name → (pattern, cve_hint))
VULN_BANNER_DB = {
    "Apache/2.2":       ("Apache/2.2",       "EOL — upgrade to 2.4+, CVE-2017-7679 etc."),
    "Apache/2.4.49":    ("Apache/2.4.49",    "Path traversal CVE-2021-41773"),
    "Apache/2.4.50":    ("Apache/2.4.50",    "Path traversal CVE-2021-42013"),
    "nginx/1.0":        ("nginx/1.0",        "EOL nginx branch"),
    "nginx/1.2":        ("nginx/1.2",        "EOL nginx branch"),
    "nginx/1.6":        ("nginx/1.6",        "EOL nginx branch"),
    "nginx/1.8":        ("nginx/1.8",        "EOL nginx branch"),
    "nginx/1.10":       ("nginx/1.10",       "EOL nginx branch"),
    "nginx/1.12":       ("nginx/1.12",       "EOL nginx branch"),
    "nginx/1.14":       ("nginx/1.14",       "EOL nginx branch"),
    "nginx/1.16":       ("nginx/1.16",       "EOL nginx branch"),
    "OpenSSL/1.0":      ("OpenSSL/1.0",      "EOL OpenSSL — multiple CVEs"),
    "OpenSSL/1.1.0":    ("OpenSSL/1.1.0",    "EOL OpenSSL 1.1.0"),
    "PHP/5":            ("PHP/5",            "EOL PHP 5 — many unpatched CVEs"),
    "PHP/7.0":          ("PHP/7.0",          "EOL PHP 7.0"),
    "PHP/7.1":          ("PHP/7.1",          "EOL PHP 7.1"),
    "PHP/7.2":          ("PHP/7.2",          "EOL PHP 7.2"),
    "Microsoft-IIS/6":  ("Microsoft-IIS/6",  "EOL IIS 6, CVE-2017-7269 WebDAV RCE"),
    "Microsoft-IIS/7":  ("Microsoft-IIS/7",  "EOL IIS 7"),
}

# Security headers that SHOULD be present
REQUIRED_SEC_HEADERS = [
    ("Strict-Transport-Security", "HSTS missing — MITM/downgrade risk"),
    ("Content-Security-Policy",   "CSP missing — XSS/injection risk"),
    ("X-Frame-Options",           "X-Frame-Options missing — clickjacking risk"),
    ("X-Content-Type-Options",    "X-Content-Type-Options missing — MIME-sniffing risk"),
    ("Referrer-Policy",           "Referrer-Policy missing — info leakage risk"),
    ("Permissions-Policy",        "Permissions-Policy missing — feature exposure risk"),
]

# Headers that should NOT be present (info leakage)
LEAK_HEADERS = ["Server", "X-Powered-By", "X-AspNet-Version",
                "X-AspNetMvc-Version", "X-Generator"]

# Common sensitive / open directories to probe
OPEN_DIR_PATHS = [
    "/.git/HEAD",        "/admin/",         "/admin/login",
    "/.env",             "/config.php",     "/wp-admin/",
    "/phpmyadmin/",      "/backup/",        "/backups/",
    "/db/",              "/.htaccess",      "/server-status",
    "/server-info",      "/elmah.axd",      "/trace.axd",
    "/actuator/",        "/actuator/env",   "/actuator/health",
    "/.DS_Store",        "/crossdomain.xml","/sitemap.xml",
    "/robots.txt",       "/.well-known/security.txt",
]

def vuln_scan(host, headers_data=None, progress_cb=None):
    """
    Runs three sub-checks:
      1. Banner / version analysis (from HTTP Server header)
      2. Missing / misconfigured security headers
      3. Open / sensitive directory probe
    Returns dict with keys: banners, headers, open_dirs, summary
    """
    import urllib.request, urllib.error

    findings = {"banners": [], "headers": [], "open_dirs": [], "summary": {}}

    # ── 1. Banner check ──────────────────────────────────────────────────────
    server_banner = ""
    if headers_data:
        for k, v in headers_data.items():
            if "server" in k.lower() or "x-powered-by" in k.lower():
                server_banner += f" {v}"
    else:
        # Fetch fresh if not provided
        try:
            req  = urllib.request.Request(
                f"https://{host}", headers={"User-Agent": "SecurityToolkit/1.0"})
            resp = urllib.request.urlopen(req, timeout=5)
            server_banner = resp.headers.get("Server", "") + " " + \
                            resp.headers.get("X-Powered-By", "")
        except Exception:
            try:
                req  = urllib.request.Request(
                    f"http://{host}", headers={"User-Agent": "SecurityToolkit/1.0"})
                resp = urllib.request.urlopen(req, timeout=5)
                server_banner = resp.headers.get("Server", "") + " " + \
                                resp.headers.get("X-Powered-By", "")
            except Exception:
                pass

    for label, (pattern, note) in VULN_BANNER_DB.items():
        if pattern.lower() in server_banner.lower():
            findings["banners"].append({
                "software": label,
                "detected": server_banner.strip(),
                "risk":     note,
                "severity": "HIGH",
            })

    if progress_cb:
        progress_cb(1)

    # ── 2. Security-header checks ─────────────────────────────────────────────
    if headers_data is None:
        headers_data = {}
        try:
            req  = urllib.request.Request(
                f"https://{host}", headers={"User-Agent": "SecurityToolkit/1.0"})
            resp = urllib.request.urlopen(req, timeout=5)
            headers_data = dict(resp.headers)
        except Exception:
            try:
                req  = urllib.request.Request(
                    f"http://{host}", headers={"User-Agent": "SecurityToolkit/1.0"})
                resp = urllib.request.urlopen(req, timeout=5)
                headers_data = dict(resp.headers)
            except Exception:
                pass

    keys_lower = {k.lower(): v for k, v in headers_data.items()}

    for hdr, note in REQUIRED_SEC_HEADERS:
        if hdr.lower() not in keys_lower:
            findings["headers"].append({
                "header":   hdr,
                "status":   "MISSING",
                "risk":     note,
                "severity": "MEDIUM",
            })
        else:
            val = keys_lower[hdr.lower()]
            # Extra checks
            if hdr == "Strict-Transport-Security" and "max-age=0" in val:
                findings["headers"].append({
                    "header":   hdr,
                    "status":   "MISCONFIGURED (max-age=0)",
                    "risk":     "HSTS disabled by max-age=0",
                    "severity": "HIGH",
                })
            elif hdr == "X-Frame-Options" and val.upper() not in ("DENY", "SAMEORIGIN"):
                findings["headers"].append({
                    "header":   hdr,
                    "status":   f"WEAK ({val})",
                    "risk":     "Non-standard X-Frame-Options value",
                    "severity": "LOW",
                })

    # Info-leakage headers
    for lk in LEAK_HEADERS:
        if lk.lower() in keys_lower:
            findings["headers"].append({
                "header":   lk,
                "status":   f"EXPOSED ({keys_lower[lk.lower()][:60]})",
                "risk":     f"{lk} header leaks server technology info",
                "severity": "INFO",
            })

    if progress_cb:
        progress_cb(2)

    # ── 3. Open-directory / sensitive-path probe ──────────────────────────────
    base_urls = [f"https://{host}", f"http://{host}"]
    # Try to detect which scheme is live
    live_base = base_urls[0]
    for b in base_urls:
        try:
            urllib.request.urlopen(
                urllib.request.Request(b, headers={"User-Agent": "SecurityToolkit/1.0"}),
                timeout=4)
            live_base = b
            break
        except urllib.error.HTTPError:
            live_base = b
            break
        except Exception:
            continue

    dir_threads = []
    dir_lock    = threading.Lock()

    def probe_path(path):
        url = live_base.rstrip("/") + path
        try:
            req  = urllib.request.Request(
                url, headers={"User-Agent": "SecurityToolkit/1.0"})
            resp = urllib.request.urlopen(req, timeout=4)
            code = resp.status
        except urllib.error.HTTPError as e:
            code = e.code
        except Exception:
            code = None

        if code in (200, 301, 302, 403):
            sev = "HIGH" if code in (200, 301, 302) else "INFO"
            risk_map = {
                200: "Accessible — possible sensitive exposure",
                301: "Redirects (may expose resource)",
                302: "Redirects (may expose resource)",
                403: "Forbidden but exists (confirms path)",
            }
            with dir_lock:
                findings["open_dirs"].append({
                    "path":     path,
                    "url":      url,
                    "code":     code,
                    "risk":     risk_map.get(code, "Unexpected response"),
                    "severity": sev,
                })

    for path in OPEN_DIR_PATHS:
        t = threading.Thread(target=probe_path, args=(path,), daemon=True)
        dir_threads.append(t)
        t.start()

    for t in dir_threads:
        t.join(timeout=6)

    if progress_cb:
        progress_cb(3)

    # ── Summary ───────────────────────────────────────────────────────────────
    all_findings = findings["banners"] + findings["headers"] + findings["open_dirs"]
    sev_count    = {"HIGH": 0, "MEDIUM": 0, "LOW": 0, "INFO": 0}
    for f in all_findings:
        sev_count[f.get("severity", "INFO")] += 1

    risk_score = sev_count["HIGH"] * 10 + sev_count["MEDIUM"] * 5 + \
                 sev_count["LOW"] * 2 + sev_count["INFO"]
    if   risk_score == 0:         level = "CLEAN"
    elif risk_score <= 10:        level = "LOW"
    elif risk_score <= 30:        level = "MEDIUM"
    elif risk_score <= 60:        level = "HIGH"
    else:                         level = "CRITICAL"

    findings["summary"] = {
        "total_findings": len(all_findings),
        "high":           sev_count["HIGH"],
        "medium":         sev_count["MEDIUM"],
        "low":            sev_count["LOW"],
        "info":           sev_count["INFO"],
        "risk_score":     risk_score,
        "risk_level":     level,
    }
    return findings


# ─── PDF Report ───────────────────────────────────────────────────────────────

def generate_pdf(report_data, output_path):
    from reportlab.lib.pagesizes import letter
    from reportlab.lib import colors
    from reportlab.lib.units import inch
    from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer,
                                    Table, TableStyle, HRFlowable, PageBreak)
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT

    doc = SimpleDocTemplate(
        output_path,
        pagesize=letter,
        leftMargin=0.75*inch,
        rightMargin=0.75*inch,
        topMargin=0.75*inch,
        bottomMargin=0.75*inch
    )

    # Styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("TitleStyle", fontSize=26, textColor=colors.HexColor("#58a6ff"),
                                  spaceAfter=6, alignment=TA_CENTER, fontName="Helvetica-Bold")
    sub_style   = ParagraphStyle("SubStyle", fontSize=11, textColor=colors.HexColor("#8b949e"),
                                  spaceAfter=4, alignment=TA_CENTER)
    h1_style    = ParagraphStyle("H1", fontSize=14, textColor=colors.HexColor("#58a6ff"),
                                  spaceBefore=14, spaceAfter=6, fontName="Helvetica-Bold")
    h2_style    = ParagraphStyle("H2", fontSize=11, textColor=colors.HexColor("#d29922"),
                                  spaceBefore=8, spaceAfter=4, fontName="Helvetica-Bold")
    body_style  = ParagraphStyle("Body", fontSize=9, textColor=colors.HexColor("#e6edf3"),
                                  leading=14, spaceAfter=3)
    dim_style   = ParagraphStyle("Dim", fontSize=8, textColor=colors.HexColor("#8b949e"),
                                  leading=12)

    def hr():
        return HRFlowable(width="100%", thickness=1,
                          color=colors.HexColor("#30363d"), spaceAfter=8, spaceBefore=4)

    def section_table(rows, col_widths=None):
        if not rows:
            return Paragraph("No data collected.", dim_style)
        if col_widths is None:
            col_widths = [2.2*inch, 4.5*inch]
        t = Table(rows, colWidths=col_widths, repeatRows=1)
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#21262d")),
            ("TEXTCOLOR",  (0, 0), (-1, 0), colors.HexColor("#58a6ff")),
            ("FONTNAME",   (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE",   (0, 0), (-1, -1), 8),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1),
             [colors.HexColor("#0d1117"), colors.HexColor("#161b22")]),
            ("TEXTCOLOR",  (0, 1), (-1, -1), colors.HexColor("#e6edf3")),
            ("GRID",       (0, 0), (-1, -1), 0.5, colors.HexColor("#30363d")),
            ("LEFTPADDING", (0, 0), (-1, -1), 6),
            ("RIGHTPADDING",(0, 0), (-1, -1), 6),
            ("TOPPADDING",  (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING",(0, 0), (-1, -1), 4),
            ("VALIGN",     (0, 0), (-1, -1), "TOP"),
            ("WORDWRAP",   (0, 0), (-1, -1), True),
        ]))
        return t

    story = []

    # ── Cover Page ──
    story.append(Spacer(1, 1.2*inch))
    story.append(Paragraph("🔐 Security Assessment Report", title_style))
    story.append(Spacer(1, 0.1*inch))
    story.append(Paragraph(f"Target: {report_data['target']}", sub_style))
    story.append(Paragraph(f"Generated: {report_data['timestamp']}", dim_style))
    story.append(Spacer(1, 0.4*inch))
    story.append(hr())

    # Summary box
    open_ports = [r for r in report_data.get("ports", []) if r[2] == "OPEN"]
    summary_rows = [
        ["Metric", "Value"],
        ["Target Host", report_data["target"]],
        ["Scan Time", report_data["timestamp"]],
        ["Ports Scanned", str(len(report_data.get("ports", [])))],
        ["Open Ports Found", str(len(open_ports))],
        ["DNS Resolution", report_data.get("dns", {}).get("IPv4", "N/A")],
    ]
    story.append(section_table(summary_rows))
    story.append(PageBreak())

    # ── Port Scan ──
    story.append(Paragraph("Port Scan Results", h1_style))
    story.append(hr())
    if report_data.get("ports"):
        port_rows = [["Port", "Service", "Status"]]
        for port, service, status in report_data["ports"]:
            port_rows.append([str(port), service, status])
        t = Table(port_rows, colWidths=[1.1*inch, 2.5*inch, 1.5*inch])
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#21262d")),
            ("TEXTCOLOR",  (0, 0), (-1, 0), colors.HexColor("#58a6ff")),
            ("FONTNAME",   (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE",   (0, 0), (-1, -1), 9),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1),
             [colors.HexColor("#0d1117"), colors.HexColor("#161b22")]),
            ("TEXTCOLOR",  (0, 1), (-1, -1), colors.HexColor("#e6edf3")),
            ("GRID",       (0, 0), (-1, -1), 0.5, colors.HexColor("#30363d")),
            ("LEFTPADDING", (0, 0), (-1, -1), 8),
            ("RIGHTPADDING",(0, 0), (-1, -1), 8),
            ("TOPPADDING",  (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING",(0, 0), (-1, -1), 5),
        ]))
        # Color-code open/closed
        for i, (port, service, status) in enumerate(report_data["ports"], start=1):
            col = colors.HexColor("#3fb950") if status == "OPEN" else colors.HexColor("#8b949e")
            t.setStyle(TableStyle([("TEXTCOLOR", (2, i), (2, i), col)]))
        story.append(t)
    else:
        story.append(Paragraph("No port scan data.", dim_style))

    story.append(PageBreak())

    # ── DNS Lookup ──
    story.append(Paragraph("DNS Lookup", h1_style))
    story.append(hr())
    dns = report_data.get("dns", {})
    if dns:
        rows = [["Field", "Value"]] + [[k, v] for k, v in dns.items()]
        story.append(section_table(rows))
    else:
        story.append(Paragraph("No DNS data collected.", dim_style))

    story.append(Spacer(1, 0.3*inch))

    # ── HTTP Headers ──
    story.append(Paragraph("HTTP Header Analysis", h1_style))
    story.append(hr())
    hdrs = report_data.get("headers", {})
    if hdrs:
        rows = [["Header", "Value"]] + [[k, str(v)[:80]] for k, v in hdrs.items()]
        story.append(section_table(rows))
        # Security header recommendations
        story.append(Spacer(1, 0.15*inch))
        story.append(Paragraph("Security Header Checklist", h2_style))
        security_headers = [
            "Strict-Transport-Security", "Content-Security-Policy",
            "X-Frame-Options", "X-Content-Type-Options",
            "Referrer-Policy", "Permissions-Policy"
        ]
        check_rows = [["Security Header", "Present?"]]
        for sh in security_headers:
            found = any(sh.lower() in k.lower() for k in hdrs)
            check_rows.append([sh, "✓ Yes" if found else "✗ Missing"])
        ct = Table(check_rows, colWidths=[3.5*inch, 1.5*inch])
        ct.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#21262d")),
            ("TEXTCOLOR",  (0, 0), (-1, 0), colors.HexColor("#58a6ff")),
            ("FONTNAME",   (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE",   (0, 0), (-1, -1), 9),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1),
             [colors.HexColor("#0d1117"), colors.HexColor("#161b22")]),
            ("TEXTCOLOR",  (0, 1), (-1, -1), colors.HexColor("#e6edf3")),
            ("GRID",       (0, 0), (-1, -1), 0.5, colors.HexColor("#30363d")),
            ("LEFTPADDING", (0, 0), (-1, -1), 8),
            ("RIGHTPADDING",(0, 0), (-1, -1), 8),
            ("TOPPADDING",  (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING",(0, 0), (-1, -1), 5),
        ]))
        for i, sh in enumerate(security_headers, start=1):
            found = any(sh.lower() in k.lower() for k in hdrs)
            col = colors.HexColor("#3fb950") if found else colors.HexColor("#f85149")
            ct.setStyle(TableStyle([("TEXTCOLOR", (1, i), (1, i), col)]))
        story.append(ct)
    else:
        story.append(Paragraph("No HTTP header data collected.", dim_style))

    story.append(PageBreak())

    # ── WHOIS ──
    story.append(Paragraph("WHOIS Information", h1_style))
    story.append(hr())
    whois = report_data.get("whois", {})
    if whois:
        rows = [["Field", "Value"]] + [[k, str(v)[:80]] for k, v in whois.items()]
        story.append(section_table(rows))
    else:
        story.append(Paragraph("No WHOIS data collected.", dim_style))

    story.append(PageBreak())

    # ── Ping / Latency ──
    story.append(Paragraph("Ping / Latency Test", h1_style))
    story.append(hr())
    ping = report_data.get("ping", {})
    if ping:
        s = ping.get("summary", ping)
        meta_rows = [["Metric", "Value"],
                     ["Host",          s.get("host", "")],
                     ["IP",            s.get("ip", "")],
                     ["Method",        s.get("method", "")],
                     ["Packets Sent",  str(s.get("sent", ""))],
                     ["Received",      str(s.get("received", ""))],
                     ["Packet Loss",   f"{s.get('loss_pct', '')}%"],
                     ["Min RTT",       f"{s.get('min_ms', 'N/A')} ms"],
                     ["Max RTT",       f"{s.get('max_ms', 'N/A')} ms"],
                     ["Avg RTT",       f"{s.get('avg_ms', 'N/A')} ms"]]
        story.append(section_table(meta_rows))
        story.append(Spacer(1, 0.15*inch))
        story.append(Paragraph("Probe Detail", h2_style))
        probe_rows = [["Seq", "Latency (ms)", "Status"]]
        for p in s.get("probes", []):
            probe_rows.append([
                str(p["seq"]),
                str(p["latency_ms"]) if p["latency_ms"] else "—",
                p["status"],
            ])
        pt = Table(probe_rows, colWidths=[0.8*inch, 2.0*inch, 2.0*inch])
        pt.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#21262d")),
            ("TEXTCOLOR",  (0, 0), (-1, 0), colors.HexColor("#58a6ff")),
            ("FONTNAME",   (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE",   (0, 0), (-1, -1), 9),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1),
             [colors.HexColor("#0d1117"), colors.HexColor("#161b22")]),
            ("TEXTCOLOR",  (0, 1), (-1, -1), colors.HexColor("#e6edf3")),
            ("GRID",       (0, 0), (-1, -1), 0.5, colors.HexColor("#30363d")),
            ("LEFTPADDING", (0, 0), (-1, -1), 8),
            ("RIGHTPADDING",(0, 0), (-1, -1), 8),
            ("TOPPADDING",  (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING",(0, 0), (-1, -1), 5),
        ]))
        for i, p in enumerate(s.get("probes", []), start=1):
            col = colors.HexColor("#3fb950") if p["status"] == "OK" else colors.HexColor("#f85149")
            pt.setStyle(TableStyle([("TEXTCOLOR", (2, i), (2, i), col)]))
        story.append(pt)
    else:
        story.append(Paragraph("No ping data collected.", dim_style))

    story.append(PageBreak())

    # ── Vulnerability Scanner ──
    story.append(Paragraph("Vulnerability Scan", h1_style))
    story.append(hr())
    vuln = report_data.get("vuln", {})
    if vuln:
        sm = vuln.get("summary", {})
        risk_color = {
            "CLEAN": "#3fb950", "LOW": "#3fb950", "MEDIUM": "#d29922",
            "HIGH": "#f85149", "CRITICAL": "#f85149"
        }.get(sm.get("risk_level", "LOW"), "#d29922")
        sum_rows = [["Metric", "Value"],
                    ["Risk Level",     sm.get("risk_level", "")],
                    ["Risk Score",     str(sm.get("risk_score", ""))],
                    ["Total Findings", str(sm.get("total_findings", ""))],
                    ["HIGH",           str(sm.get("high", 0))],
                    ["MEDIUM",         str(sm.get("medium", 0))],
                    ["LOW",            str(sm.get("low", 0))],
                    ["INFO",           str(sm.get("info", 0))]]
        st = section_table(sum_rows)
        story.append(st)

        def vuln_table(title, rows_data, cols, widths):
            if not rows_data:
                story.append(Spacer(1, 0.1*inch))
                story.append(Paragraph(f"{title}: none found.", dim_style))
                return
            story.append(Spacer(1, 0.15*inch))
            story.append(Paragraph(title, h2_style))
            tbl_rows = [cols]
            for r in rows_data:
                tbl_rows.append([str(r.get(c.lower().replace(" ", "_"), "")) for c in cols])
            t = Table(tbl_rows, colWidths=widths)
            sev_colors = {"HIGH": "#f85149", "MEDIUM": "#d29922",
                          "LOW": "#3fb950",  "INFO": "#8b949e"}
            t.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#21262d")),
                ("TEXTCOLOR",  (0, 0), (-1, 0), colors.HexColor("#58a6ff")),
                ("FONTNAME",   (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE",   (0, 0), (-1, -1), 8),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1),
                 [colors.HexColor("#0d1117"), colors.HexColor("#161b22")]),
                ("TEXTCOLOR",  (0, 1), (-1, -1), colors.HexColor("#e6edf3")),
                ("GRID",       (0, 0), (-1, -1), 0.5, colors.HexColor("#30363d")),
                ("LEFTPADDING", (0, 0), (-1, -1), 6),
                ("RIGHTPADDING",(0, 0), (-1, -1), 6),
                ("TOPPADDING",  (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING",(0, 0), (-1, -1), 4),
                ("WORDWRAP",   (0, 0), (-1, -1), True),
            ]))
            # Color severity column (last col assumed)
            sc = len(cols) - 1
            for i, r in enumerate(rows_data, start=1):
                sev = r.get("severity", "INFO")
                t.setStyle(TableStyle([
                    ("TEXTCOLOR", (sc, i), (sc, i),
                     colors.HexColor(sev_colors.get(sev, "#8b949e")))
                ]))
            story.append(t)

        vuln_table("Outdated / Vulnerable Software",
                   vuln.get("banners", []),
                   ["Software", "Risk", "Severity"],
                   [1.5*inch, 4.2*inch, 0.9*inch])

        vuln_table("Security Header Issues",
                   vuln.get("headers", []),
                   ["Header", "Status", "Severity"],
                   [2.0*inch, 3.5*inch, 0.9*inch])

        vuln_table("Open / Sensitive Directories",
                   vuln.get("open_dirs", []),
                   ["Path", "Code", "Severity"],
                   [3.0*inch, 0.7*inch, 0.9*inch])
    else:
        story.append(Paragraph("No vulnerability scan data collected.", dim_style))

    story.append(Spacer(1, 0.5*inch))
    story.append(hr())
    story.append(Paragraph(
        "Generated by Security Assessment Toolkit  |  For authorized use only",
        ParagraphStyle("Footer", fontSize=7, textColor=colors.HexColor("#8b949e"),
                       alignment=TA_CENTER)
    ))

    doc.build(story)


# ─── Excel Export ─────────────────────────────────────────────────────────────

def generate_excel(report_data, output_path):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    hdr_font  = Font(name="Calibri", bold=True, color="58A6FF", size=10)
    hdr_fill  = PatternFill("solid", fgColor="21262D")
    row_fills = [PatternFill("solid", fgColor="0D1117"),
                 PatternFill("solid", fgColor="161B22")]
    cell_font = Font(name="Calibri", color="E6EDF3", size=9)
    thin      = Side(style="thin", color="30363D")
    border    = Border(left=thin, right=thin, top=thin, bottom=thin)
    center    = Alignment(horizontal="center", vertical="center")

    def style_header(ws, row_idx, cols):
        for c in range(1, cols + 1):
            cell = ws.cell(row=row_idx, column=c)
            cell.font  = hdr_font
            cell.fill  = hdr_fill
            cell.border = border
            cell.alignment = center

    def style_data_row(ws, row_idx, cols, alt):
        for c in range(1, cols + 1):
            cell = ws.cell(row=row_idx, column=c)
            cell.font   = cell_font
            cell.fill   = row_fills[alt % 2]
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    def autofit(ws, padding=4):
        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=0)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + padding, 60)

    # ── Summary Sheet ──
    ws = wb.active
    ws.title = "Summary"
    ws.sheet_properties.tabColor = "58A6FF"
    ws.append(["Security Assessment Summary"])
    ws["A1"].font = Font(name="Calibri", bold=True, size=14, color="58A6FF")
    ws.append([])
    headers = ["Metric", "Value"]
    ws.append(headers)
    style_header(ws, 3, 2)
    open_ports = [r for r in report_data.get("ports", []) if r[2] == "OPEN"]
    summary_rows = [
        ("Target Host",     report_data["target"]),
        ("Scan Time",       report_data["timestamp"]),
        ("Ports Scanned",   len(report_data.get("ports", []))),
        ("Open Ports",      len(open_ports)),
        ("DNS IPv4",        report_data.get("dns", {}).get("IPv4", "N/A")),
        ("DNS IPv6",        report_data.get("dns", {}).get("IPv6", "N/A")),
    ]
    for i, (k, v) in enumerate(summary_rows, start=4):
        ws.cell(row=i, column=1).value = k
        ws.cell(row=i, column=2).value = str(v)
        style_data_row(ws, i, 2, i)
    autofit(ws)

    # ── Port Scan Sheet ──
    ws2 = wb.create_sheet("Port Scan")
    ws2.sheet_properties.tabColor = "3FB950"
    ws2.append(["Port", "Service", "Status"])
    style_header(ws2, 1, 3)
    for i, (port, service, status) in enumerate(report_data.get("ports", []), start=2):
        ws2.cell(row=i, column=1).value = port
        ws2.cell(row=i, column=2).value = service
        ws2.cell(row=i, column=3).value = status
        style_data_row(ws2, i, 3, i)
        color = "3FB950" if status == "OPEN" else "8B949E"
        ws2.cell(row=i, column=3).font = Font(name="Calibri", color=color, bold=(status=="OPEN"), size=9)
    autofit(ws2)

    # ── DNS Sheet ──
    ws3 = wb.create_sheet("DNS Lookup")
    ws3.sheet_properties.tabColor = "D29922"
    ws3.append(["Field", "Value"])
    style_header(ws3, 1, 2)
    for i, (k, v) in enumerate(report_data.get("dns", {}).items(), start=2):
        ws3.cell(row=i, column=1).value = k
        ws3.cell(row=i, column=2).value = str(v)
        style_data_row(ws3, i, 2, i)
    autofit(ws3)

    # ── Headers Sheet ──
    ws4 = wb.create_sheet("HTTP Headers")
    ws4.sheet_properties.tabColor = "388BFD"
    ws4.append(["Header", "Value"])
    style_header(ws4, 1, 2)
    for i, (k, v) in enumerate(report_data.get("headers", {}).items(), start=2):
        ws4.cell(row=i, column=1).value = k
        ws4.cell(row=i, column=2).value = str(v)
        style_data_row(ws4, i, 2, i)
    autofit(ws4)

    # ── WHOIS Sheet ──
    ws5 = wb.create_sheet("WHOIS")
    ws5.sheet_properties.tabColor = "F85149"
    ws5.append(["Field", "Value"])
    style_header(ws5, 1, 2)
    for i, (k, v) in enumerate(report_data.get("whois", {}).items(), start=2):
        ws5.cell(row=i, column=1).value = k
        ws5.cell(row=i, column=2).value = str(v)
        style_data_row(ws5, i, 2, i)
    autofit(ws5)

    # ── Ping Sheet ──
    ws6 = wb.create_sheet("Ping")
    ws6.sheet_properties.tabColor = "3FB950"
    ping = report_data.get("ping", {})
    s = ping.get("summary", ping) if ping else {}
    ws6.append(["Metric", "Value"])
    style_header(ws6, 1, 2)
    ping_meta = [
        ("Host",          s.get("host", "")),
        ("IP",            s.get("ip", "")),
        ("Method",        s.get("method", "")),
        ("Sent",          s.get("sent", "")),
        ("Received",      s.get("received", "")),
        ("Lost",          s.get("lost", "")),
        ("Loss %",        f"{s.get('loss_pct', '')}%"),
        ("Min RTT (ms)",  s.get("min_ms", "N/A")),
        ("Max RTT (ms)",  s.get("max_ms", "N/A")),
        ("Avg RTT (ms)",  s.get("avg_ms", "N/A")),
    ]
    for i, (k, v) in enumerate(ping_meta, start=2):
        ws6.cell(row=i, column=1).value = k
        ws6.cell(row=i, column=2).value = str(v)
        style_data_row(ws6, i, 2, i)
    # Probe detail below
    start_row = len(ping_meta) + 3
    ws6.cell(row=start_row, column=1).value = "Seq"
    ws6.cell(row=start_row, column=2).value = "Latency (ms)"
    ws6.cell(row=start_row, column=3).value = "Status"
    style_header(ws6, start_row, 3)
    for j, p in enumerate(s.get("probes", []), start=start_row + 1):
        ws6.cell(row=j, column=1).value = p["seq"]
        ws6.cell(row=j, column=2).value = p["latency_ms"] if p["latency_ms"] else "—"
        ws6.cell(row=j, column=3).value = p["status"]
        style_data_row(ws6, j, 3, j)
        col = "3FB950" if p["status"] == "OK" else "F85149"
        ws6.cell(row=j, column=3).font = Font(name="Calibri", color=col, size=9, bold=True)
    autofit(ws6)

    # ── Vulnerability Sheet ──
    ws7 = wb.create_sheet("Vuln Scan")
    ws7.sheet_properties.tabColor = "F85149"
    vuln = report_data.get("vuln", {})
    sev_color_map = {"HIGH": "F85149", "MEDIUM": "D29922", "LOW": "3FB950", "INFO": "8B949E"}

    row_cursor = 1
    # Summary block
    sm = vuln.get("summary", {})
    ws7.cell(row=row_cursor, column=1).value = "Risk Level"
    ws7.cell(row=row_cursor, column=2).value = sm.get("risk_level", "N/A")
    ws7.cell(row=row_cursor, column=3).value = "Risk Score"
    ws7.cell(row=row_cursor, column=4).value = sm.get("risk_score", 0)
    for c in range(1, 5):
        ws7.cell(row=row_cursor, column=c).font = Font(name="Calibri", bold=True, color="58A6FF", size=10)
        ws7.cell(row=row_cursor, column=c).fill = hdr_fill
        ws7.cell(row=row_cursor, column=c).border = border
    row_cursor += 2

    def vuln_section(title, items, cols):
        nonlocal row_cursor
        ws7.cell(row=row_cursor, column=1).value = title
        ws7.cell(row=row_cursor, column=1).font = Font(name="Calibri", bold=True, color="D29922", size=10)
        row_cursor += 1
        for ci, col in enumerate(cols, 1):
            ws7.cell(row=row_cursor, column=ci).value = col
        style_header(ws7, row_cursor, len(cols))
        row_cursor += 1
        for item in items:
            for ci, col in enumerate(cols, 1):
                ws7.cell(row=row_cursor, column=ci).value = str(item.get(col.lower().replace(" ", "_"), ""))
            style_data_row(ws7, row_cursor, len(cols), row_cursor)
            sev = item.get("severity", "INFO")
            sev_col_idx = len(cols)
            ws7.cell(row=row_cursor, column=sev_col_idx).font = Font(
                name="Calibri", color=sev_color_map.get(sev, "8B949E"), size=9, bold=True)
            row_cursor += 1
        row_cursor += 1

    vuln_section("Outdated / Vulnerable Software",
                 vuln.get("banners", []),  ["Software", "Risk", "Severity"])
    vuln_section("Security Header Issues",
                 vuln.get("headers", []),  ["Header",   "Status", "Risk", "Severity"])
    vuln_section("Open / Sensitive Directories",
                 vuln.get("open_dirs", []),["Path",     "URL",  "Code", "Severity"])
    autofit(ws7)

    wb.save(output_path)


# ─── GUI ──────────────────────────────────────────────────────────────────────

class SecurityToolkit(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("🔐 Security Assessment Toolkit")
        self.configure(bg=BG)
        self.geometry("950x780")
        self.minsize(800, 600)
        self.resizable(True, True)

        self._report_data = {}
        self._scan_running = False
        self._total_tasks  = 0
        self._done_tasks   = 0

        self._setup_styles()
        self._build_ui()

    def _setup_styles(self):
        style = ttk.Style(self)
        style.theme_use("default")
        style.configure("Dark.TFrame",      background=BG)
        style.configure("Card.TFrame",      background=BG2, relief="flat")
        style.configure("Dark.TLabel",      background=BG,  foreground=TEXT,     font=("Segoe UI", 10))
        style.configure("Card.TLabel",      background=BG2, foreground=TEXT,     font=("Segoe UI", 10))
        style.configure("Title.TLabel",     background=BG,  foreground=ACCENT,   font=("Segoe UI", 20, "bold"))
        style.configure("Sub.TLabel",       background=BG,  foreground=TEXT_DIM, font=("Segoe UI", 9))
        style.configure("Accent.TButton",   background=ACCENT2, foreground="white",
                        font=("Segoe UI", 10, "bold"), relief="flat", padding=(12, 6))
        style.configure("Danger.TButton",   background="#b91c1c", foreground="white",
                        font=("Segoe UI", 10, "bold"), relief="flat", padding=(12, 6))
        style.configure("Dark.TEntry",      fieldbackground=BG3, foreground=TEXT,
                        insertcolor=TEXT, font=("Segoe UI", 11))
        style.configure("Dark.TCheckbutton",background=BG2, foreground=TEXT,     font=("Segoe UI", 9))
        style.configure("Dark.Horizontal.TProgressbar",
                        troughcolor=BG3, background=ACCENT, thickness=6)
        style.configure("Dark.TNotebook",   background=BG,  tabmargins=[2, 2, 0, 0])
        style.configure("Dark.TNotebook.Tab",
                        background=BG3, foreground=TEXT_DIM, padding=[12, 5],
                        font=("Segoe UI", 9))
        style.map("Dark.TNotebook.Tab",
                  background=[("selected", BG2)],
                  foreground=[("selected", ACCENT)])

    # ── UI Construction ──

    def _build_ui(self):
        # Header
        hdr = tk.Frame(self, bg=BG2, bd=0)
        hdr.pack(fill="x", padx=0, pady=0)
        tk.Label(hdr, text="🔐 Security Assessment Toolkit", bg=BG2, fg=ACCENT,
                 font=("Segoe UI", 16, "bold")).pack(side="left", padx=20, pady=12)
        tk.Label(hdr, text="Authorized use only", bg=BG2, fg=TEXT_DIM,
                 font=("Segoe UI", 8)).pack(side="right", padx=20)

        # Target row
        target_frame = tk.Frame(self, bg=BG, pady=12)
        target_frame.pack(fill="x", padx=20)
        tk.Label(target_frame, text="Target Host:", bg=BG, fg=TEXT_DIM,
                 font=("Segoe UI", 9)).pack(side="left")
        self.target_var = tk.StringVar(value="scanme.nmap.org")
        entry = tk.Entry(target_frame, textvariable=self.target_var, bg=BG3, fg=TEXT,
                         insertbackground=TEXT, font=("Segoe UI", 11), relief="flat",
                         bd=0, width=35)
        entry.pack(side="left", padx=(8, 16), ipady=6)
        entry.bind("<Return>", lambda e: self._start_scan())

        self.scan_btn = tk.Button(target_frame, text="▶  Start Scan", bg=ACCENT2, fg="white",
                                   font=("Segoe UI", 10, "bold"), relief="flat", cursor="hand2",
                                   padx=16, pady=6, command=self._start_scan)
        self.scan_btn.pack(side="left", padx=(0, 8))

        self.stop_btn = tk.Button(target_frame, text="■  Stop", bg="#b91c1c", fg="white",
                                   font=("Segoe UI", 10, "bold"), relief="flat", cursor="hand2",
                                   padx=14, pady=6, command=self._stop_scan, state="disabled")
        self.stop_btn.pack(side="left")

        # Scan Options
        opts_frame = tk.Frame(self, bg=BG2, bd=0)
        opts_frame.pack(fill="x", padx=20, pady=(0, 8))
        tk.Label(opts_frame, text="Modules:", bg=BG2, fg=TEXT_DIM,
                 font=("Segoe UI", 9, "bold")).pack(side="left", padx=(10, 8), pady=8)
        self.opt_ports   = self._checkbox(opts_frame, "Port Scan", True)
        self.opt_dns     = self._checkbox(opts_frame, "DNS Lookup", True)
        self.opt_headers = self._checkbox(opts_frame, "HTTP Headers", True)
        self.opt_whois   = self._checkbox(opts_frame, "WHOIS", True)
        self.opt_ping    = self._checkbox(opts_frame, "Ping", True)
        self.opt_vuln    = self._checkbox(opts_frame, "Vuln Scan", True)

        # Progress bar
        prog_frame = tk.Frame(self, bg=BG)
        prog_frame.pack(fill="x", padx=20, pady=(0, 6))
        self.progress_var = tk.DoubleVar(value=0)
        self.progress_bar = ttk.Progressbar(prog_frame, variable=self.progress_var,
                                            maximum=100, style="Dark.Horizontal.TProgressbar")
        self.progress_bar.pack(fill="x", side="left", expand=True)
        self.status_lbl = tk.Label(prog_frame, text="Ready", bg=BG, fg=TEXT_DIM,
                                    font=("Segoe UI", 8), width=28, anchor="e")
        self.status_lbl.pack(side="right", padx=(8, 0))

        # Notebook
        nb = ttk.Notebook(self, style="Dark.TNotebook")
        nb.pack(fill="both", expand=True, padx=20, pady=(0, 8))

        def tab(label):
            f = tk.Frame(nb, bg=BG2)
            nb.add(f, text=label)
            return f

        self.tabs = {
            "ports":   tab("⚡ Port Scan"),
            "dns":     tab("🌐 DNS"),
            "headers": tab("📋 HTTP Headers"),
            "whois":   tab("📄 WHOIS"),
            "ping":    tab("📡 Ping"),
            "vuln":    tab("🛡 Vuln Scan"),
            "log":     tab("📝 Log"),
        }
        self.outputs = {}
        for key, frame in self.tabs.items():
            st = scrolledtext.ScrolledText(
                frame, bg=BG, fg=TEXT, font=("Consolas", 9),
                insertbackground=TEXT, relief="flat", bd=0,
                selectbackground=ACCENT2, wrap="word"
            )
            st.pack(fill="both", expand=True, padx=4, pady=4)
            st.config(state="disabled")
            self.outputs[key] = st

        # Color tags
        for key in self.outputs:
            w = self.outputs[key]
            w.tag_config("open",     foreground=GREEN)
            w.tag_config("closed",   foreground=TEXT_DIM)
            w.tag_config("info",     foreground=ACCENT)
            w.tag_config("warn",     foreground=YELLOW)
            w.tag_config("error",    foreground=RED)
            w.tag_config("heading",  foreground=ACCENT, font=("Consolas", 10, "bold"))
            w.tag_config("high",     foreground=RED)
            w.tag_config("medium",   foreground=YELLOW)
            w.tag_config("low",      foreground=GREEN)
            w.tag_config("infoleak", foreground=TEXT_DIM)

        # Export buttons
        exp_frame = tk.Frame(self, bg=BG)
        exp_frame.pack(fill="x", padx=20, pady=(0, 12))
        btns = [
            ("📄 Export PDF",   self._export_pdf),
            ("📊 Export Excel", self._export_excel),
            ("📁 Export CSV",   self._export_csv),
            ("🗑  Clear",       self._clear_all),
        ]
        for label, cmd in btns:
            tk.Button(exp_frame, text=label, bg=BG3, fg=TEXT,
                      font=("Segoe UI", 9), relief="flat", cursor="hand2",
                      padx=14, pady=5, command=cmd,
                      activebackground=BG2, activeforeground=ACCENT).pack(side="left", padx=(0, 6))

    def _checkbox(self, parent, text, default=True):
        var = tk.BooleanVar(value=default)
        cb = tk.Checkbutton(parent, text=text, variable=var, bg=BG2, fg=TEXT,
                            selectcolor=BG3, activebackground=BG2,
                            font=("Segoe UI", 9), cursor="hand2")
        cb.pack(side="left", padx=(0, 12), pady=8)
        return var

    # ── Write helpers ──

    def _write(self, key, text, tag=None):
        w = self.outputs[key]
        w.config(state="normal")
        if tag:
            w.insert("end", text, tag)
        else:
            w.insert("end", text)
        w.see("end")
        w.config(state="disabled")

    def _log(self, msg, level="info"):
        ts = datetime.datetime.now().strftime("%H:%M:%S")
        tag = {"info": "info", "warn": "warn", "error": "error"}.get(level, None)
        self._write("log", f"[{ts}] {msg}\n", tag)

    def _set_status(self, msg):
        self.status_lbl.config(text=msg)
        self.update_idletasks()

    def _advance(self):
        self._done_tasks += 1
        pct = (self._done_tasks / max(self._total_tasks, 1)) * 100
        self.progress_var.set(pct)
        self.update_idletasks()

    # ── Scan Logic ──

    def _start_scan(self):
        if self._scan_running:
            return
        target = self.target_var.get().strip()
        if not target:
            messagebox.showwarning("No Target", "Please enter a target host.")
            return
        self._scan_running = True
        self.scan_btn.config(state="disabled")
        self.stop_btn.config(state="normal")
        self.progress_var.set(0)
        self._clear_all(silent=True)
        self._report_data = {
            "target": target,
            "timestamp": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "ports": [], "dns": {}, "headers": {}, "whois": {},
            "ping": {}, "vuln": {}
        }
        tasks = []
        if self.opt_ports.get():   tasks.append("ports")
        if self.opt_dns.get():     tasks.append("dns")
        if self.opt_headers.get(): tasks.append("headers")
        if self.opt_whois.get():   tasks.append("whois")
        if self.opt_ping.get():    tasks.append("ping")
        if self.opt_vuln.get():    tasks.append("vuln")
        self._total_tasks = len(tasks) * 5  # rough units
        self._done_tasks  = 0
        threading.Thread(target=self._run_scan, args=(target, tasks), daemon=True).start()

    def _stop_scan(self):
        self._scan_running = False
        self._set_status("Stopped.")
        self._log("Scan stopped by user.", "warn")

    def _run_scan(self, target, tasks):
        try:
            if "ports" in tasks and self._scan_running:
                self._scan_ports(target)
            if "dns" in tasks and self._scan_running:
                self._scan_dns(target)
            if "headers" in tasks and self._scan_running:
                self._scan_headers(target)
            if "whois" in tasks and self._scan_running:
                self._scan_whois(target)
            if "ping" in tasks and self._scan_running:
                self._scan_ping(target)
            if "vuln" in tasks and self._scan_running:
                self._scan_vuln(target)
        except Exception as e:
            self._log(f"Scan error: {e}", "error")
        finally:
            self._scan_running = False
            self.after(0, self._scan_complete)

    def _scan_ports(self, target):
        self._set_status("Scanning ports...")
        self._log(f"Starting port scan on {target}")
        self._write("ports", f"Port scan: {target}\n", "heading")
        self._write("ports", f"{'Port':<8} {'Service':<16} {'Status'}\n", "info")
        self._write("ports", "─" * 40 + "\n")

        port_threads = []
        results = []
        lock = threading.Lock()

        def check(port, service):
            if not self._scan_running:
                return
            open_ = scan_port(target, port, timeout=0.5)
            status = "OPEN" if open_ else "CLOSED"
            with lock:
                results.append((port, service, status))
            tag = "open" if open_ else "closed"
            self.after(0, lambda p=port, svc=service, s=status, t=tag:
                       self._write("ports", f"{p:<8} {svc:<16} {s}\n", t))
            self._advance()

        for port, service in COMMON_PORTS.items():
            t = threading.Thread(target=check, args=(port, service), daemon=True)
            port_threads.append(t)
            t.start()

        for t in port_threads:
            t.join(timeout=3)

        results.sort(key=lambda x: x[0])
        self._report_data["ports"] = results
        open_count = sum(1 for _, _, s in results if s == "OPEN")
        self._write("ports", f"\n✓ Done — {open_count} open port(s) found\n", "info")
        self._log(f"Port scan complete: {open_count} open port(s)")

    def _scan_dns(self, target):
        self._set_status("DNS lookup...")
        self._log(f"DNS lookup: {target}")
        self._write("dns", f"DNS Lookup: {target}\n", "heading")
        self._write("dns", "─" * 40 + "\n")
        data = dns_lookup(target)
        self._report_data["dns"] = data
        for k, v in data.items():
            self._write("dns", f"  {k:<16} ", "info")
            self._write("dns", f"{v}\n")
        self._advance()
        self._log("DNS lookup complete.")

    def _scan_headers(self, target):
        self._set_status("Fetching HTTP headers...")
        self._log(f"HTTP headers: {target}")
        self._write("headers", f"HTTP Headers: {target}\n", "heading")
        self._write("headers", "─" * 40 + "\n")
        data = http_headers(target)
        self._report_data["headers"] = data
        security_keys = {"strict-transport-security", "content-security-policy",
                         "x-frame-options", "x-content-type-options",
                         "referrer-policy", "permissions-policy"}
        for k, v in data.items():
            tag = "open" if any(sk in k.lower() for sk in security_keys) else None
            self._write("headers", f"  {k}: ", "info")
            self._write("headers", f"{v[:80]}\n", tag)
        self._advance()
        self._log("Header analysis complete.")

    def _scan_whois(self, target):
        self._set_status("WHOIS lookup...")
        self._log(f"WHOIS: {target}")
        self._write("whois", f"WHOIS: {target}\n", "heading")
        self._write("whois", "─" * 40 + "\n")
        data = whois_lookup(target)
        self._report_data["whois"] = data
        for k, v in data.items():
            self._write("whois", f"  {k:<20} ", "info")
            self._write("whois", f"{v}\n")
        self._advance()
        self._log("WHOIS complete.")

    def _scan_ping(self, target):
        self._set_status("Pinging host...")
        self._log(f"Ping: {target}")
        self._write("ping", f"Ping  →  {target}\n", "heading")
        self._write("ping", "─" * 48 + "\n")

        result = ping_host(target, count=5)
        self._report_data["ping"] = result

        s = result
        self._write("ping", f"  IP      : ", "info")
        self._write("ping", f"{s.get('ip', '?')}\n")
        self._write("ping", f"  Method  : ", "info")
        self._write("ping", f"{s.get('method', '?')}\n")
        self._write("ping", "\n")
        self._write("ping", f"  {'Seq':<6} {'Latency':>12}    {'Status'}\n", "info")
        self._write("ping", "  " + "─" * 36 + "\n")

        for p in s.get("probes", []):
            lat_str = f"{p['latency_ms']:>8.2f} ms" if p["latency_ms"] else "      timeout"
            tag = "open" if p["status"] == "OK" else "error"
            self._write("ping", f"  [{p['seq']:>2}]  {lat_str}    {p['status']}\n", tag)

        self._write("ping", "\n")
        lost  = s.get("lost",     0)
        loss  = s.get("loss_pct", 0)
        avg   = s.get("avg_ms",   None)
        reachable = s.get("received", 0) > 0

        self._write("ping", f"  Packets   sent={s.get('sent',0)}  "
                            f"received={s.get('received',0)}  "
                            f"lost={lost} ({loss}%)\n",
                    "open" if reachable else "error")
        if avg:
            self._write("ping", f"  RTT       min={s.get('min_ms')} ms  "
                                f"avg={avg} ms  max={s.get('max_ms')} ms\n", "info")

        status_str = "Reachable" if reachable else "Unreachable"
        tag        = "open" if reachable else "error"
        self._write("ping", f"\n  Host status: {status_str}\n", tag)
        self._log(f"Ping complete — {status_str}, avg {avg} ms")
        self._advance()

    def _scan_vuln(self, target):
        self._set_status("Running vulnerability scan...")
        self._log(f"Vuln scan: {target}")
        self._write("vuln", f"Vulnerability Scan  →  {target}\n", "heading")
        self._write("vuln", "─" * 56 + "\n\n")

        def progress_cb(step):
            labels = {1: "Checking banners...",
                      2: "Auditing security headers...",
                      3: "Probing sensitive directories..."}
            self._set_status(labels.get(step, "Vuln scan..."))
            self._advance()

        # Reuse already-fetched headers if available, else None
        existing_headers = self._report_data.get("headers") or None
        result = vuln_scan(target, headers_data=existing_headers, progress_cb=progress_cb)
        self._report_data["vuln"] = result

        sm = result.get("summary", {})
        risk_tag = {"CLEAN": "low", "LOW": "low", "MEDIUM": "medium",
                    "HIGH": "high", "CRITICAL": "high"}.get(sm.get("risk_level",""), "medium")

        self._write("vuln", f"  Risk Level : ", "info")
        self._write("vuln", f"{sm.get('risk_level','?')}  (score {sm.get('risk_score',0)})\n", risk_tag)
        self._write("vuln", f"  Findings   : HIGH={sm.get('high',0)}  "
                            f"MEDIUM={sm.get('medium',0)}  LOW={sm.get('low',0)}  "
                            f"INFO={sm.get('info',0)}\n\n")

        # ── Banners ──
        banners = result.get("banners", [])
        self._write("vuln", f"  ▸ Outdated / Vulnerable Software  ({len(banners)} finding(s))\n", "info")
        if banners:
            for b in banners:
                self._write("vuln", f"    [{b['severity']}] {b['software']}\n",
                            "high" if b["severity"] == "HIGH" else "medium")
                self._write("vuln", f"          {b['risk']}\n", "infoleak")
        else:
            self._write("vuln", "    ✓ No vulnerable banners detected\n", "low")

        self._write("vuln", "\n")

        # ── Headers ──
        hdr_issues = result.get("headers", [])
        self._write("vuln", f"  ▸ Security Header Issues  ({len(hdr_issues)} finding(s))\n", "info")
        if hdr_issues:
            for h in hdr_issues:
                sev_tag = {"HIGH": "high", "MEDIUM": "medium",
                           "LOW": "low", "INFO": "infoleak"}.get(h["severity"], "infoleak")
                self._write("vuln", f"    [{h['severity']}] {h['header']}: {h['status']}\n", sev_tag)
                self._write("vuln", f"          {h['risk']}\n", "infoleak")
        else:
            self._write("vuln", "    ✓ All required security headers present\n", "low")

        self._write("vuln", "\n")

        # ── Open dirs ──
        dirs = result.get("open_dirs", [])
        self._write("vuln", f"  ▸ Open / Sensitive Directories  ({len(dirs)} finding(s))\n", "info")
        if dirs:
            for d in dirs:
                sev_tag = "high" if d["severity"] == "HIGH" else "infoleak"
                self._write("vuln", f"    [{d['severity']}] HTTP {d['code']}  {d['path']}\n", sev_tag)
                self._write("vuln", f"          {d['risk']}\n", "infoleak")
        else:
            self._write("vuln", "    ✓ No sensitive paths exposed\n", "low")

        self._log(f"Vuln scan complete — risk level: {sm.get('risk_level','?')}")

    def _scan_complete(self):
        self.progress_var.set(100)
        self._set_status("Scan complete ✓")
        self.scan_btn.config(state="normal")
        self.stop_btn.config(state="disabled")
        self._log("All scans complete.", "info")

    # ── Exports ──

    def _ensure_data(self):
        if not self._report_data.get("target"):
            messagebox.showinfo("No Data", "Run a scan first.")
            return False
        return True

    def _export_pdf(self):
        if not self._ensure_data():
            return
        ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        target_safe = self._report_data["target"].replace(".", "_").replace("/", "_")
        path = os.path.join(os.path.expanduser("~"), "Desktop",
                            f"security_report_{target_safe}_{ts}.pdf")
        try:
            generate_pdf(self._report_data, path)
            self._log(f"PDF saved: {path}", "info")
            messagebox.showinfo("PDF Exported", f"Report saved to:\n{path}")
            if sys.platform == "win32":
                os.startfile(path)
            elif sys.platform == "darwin":
                os.system(f'open "{path}"')
        except Exception as e:
            self._log(f"PDF error: {e}", "error")
            messagebox.showerror("PDF Error", str(e))

    def _export_excel(self):
        if not self._ensure_data():
            return
        ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        target_safe = self._report_data["target"].replace(".", "_").replace("/", "_")
        path = os.path.join(os.pa   th.expanduser("~"), "Desktop",
                            f"security_report_{target_safe}_{ts}.xlsx")
        try:
            generate_excel(self._report_data, path)
            self._log(f"Excel saved: {path}", "info")
            messagebox.showinfo("Excel Exported", f"Report saved to:\n{path}")
        except Exception as e:
            self._log(f"Excel error: {e}", "error")
            messagebox.showerror("Excel Error", str(e))

    def _export_csv(self):
        if not self._ensure_data():
            return
        ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        target_safe = self._report_data["target"].replace(".", "_").replace("/", "_")
        path = os.path.join(os.path.expanduser("~"), "Desktop",
                            f"security_report_{target_safe}_{ts}.csv")
        try:
            with open(path, "w", newline="") as f:
                w = csv.writer(f)
                w.writerow(["Section", "Field", "Value"])
                for port, service, status in self._report_data.get("ports", []):
                    w.writerow(["Port Scan", f"{port}/{service}", status])
                for k, v in self._report_data.get("dns", {}).items():
                    w.writerow(["DNS", k, v])
                for k, v in self._report_data.get("headers", {}).items():
                    w.writerow(["HTTP Headers", k, v])
                for k, v in self._report_data.get("whois", {}).items():
                    w.writerow(["WHOIS", k, v])
                # Ping
                ping = self._report_data.get("ping", {})
                s = ping.get("summary", ping) if ping else {}
                for field in ["host","ip","method","sent","received","lost","loss_pct","min_ms","avg_ms","max_ms"]:
                    if field in s:
                        w.writerow(["Ping", field, s[field]])
                for p in s.get("probes", []):
                    w.writerow(["Ping Probe", f"seq={p['seq']}", f"{p['latency_ms']} ms  {p['status']}"])
                # Vuln
                vuln = self._report_data.get("vuln", {})
                for b in vuln.get("banners", []):
                    w.writerow(["Vuln-Banner", b.get("software",""), f"[{b.get('severity','')}] {b.get('risk','')}"])
                for h in vuln.get("headers", []):
                    w.writerow(["Vuln-Header", h.get("header",""), f"[{h.get('severity','')}] {h.get('status','')} — {h.get('risk','')}"])
                for d in vuln.get("open_dirs", []):
                    w.writerow(["Vuln-Dir", d.get("path",""), f"[{d.get('severity','')}] HTTP {d.get('code','')} — {d.get('risk','')}"])
            self._log(f"CSV saved: {path}", "info")
            messagebox.showinfo("CSV Exported", f"Report saved to:\n{path}")
        except Exception as e:
            self._log(f"CSV error: {e}", "error")
            messagebox.showerror("CSV Error", str(e))

    def _clear_all(self, silent=False):
        for key, w in self.outputs.items():
            w.config(state="normal")
            w.delete("1.0", "end")
            w.config(state="disabled")
        self.progress_var.set(0)
        self._set_status("Ready")
        if not silent:
            self._report_data = {}


if __name__ == "__main__":
    app = SecurityToolkit()
    app.mainloop()
